import typer
import pandas as pd
from utils.db import DBManager
from utils.get_info import get_finviz_data, get_quarter_medians_and_year_avg
from openpyxl import load_workbook

app = typer.Typer()

DB_PATH = "stock_db.db"

@app.command()
def download(stocks: str, period: str = "5y", db_path: str = DB_PATH):
    """
    下載指定股票們的資料並存入資料庫。以逗號分隔。例如："2330.TW,2412.TW,00675L.TW,GOOG"

    Args:
        stocks (str): 股票代碼，以逗號分隔。例如："2330.TW,2412.TW,00675L.TW,GOOG"
        period (str, optional): 資料下載期間。預設為 "5y"。
        db_path (str, optional): 資料庫路徑。預設為 "stock_db.db"。

    """
    stock_list = [s.strip() for s in stocks.split(',')]
    print(f"[CLI] 開始下載股票資料: {stock_list}，期間: {period}")
    with DBManager(db_path) as db_manager:
        for stock in stock_list:
            df = db_manager.get_stock(stock, period=period)
            if df.empty:
                print(f"[CLI] {stock} 的資料為空，請檢查下載是否成功。")
            else:
                print(f"[CLI] {stock} 的資料已載入，包含 {len(df)} 筆資料。")
    print("下載完成。")

@app.command()
def update(stocks: str, db_path: str = DB_PATH):
    """
    更新指定股票的資料，並將最新資料添加至資料庫。以逗號分隔。例如："2330.TW,2412.TW,00675L.TW,GOOG"

    Args:
        stocks (str): 股票代碼，以逗號分隔。例如："2330.TW,2412.TW,00675L.TW, GOOG"
        db_path (str, optional): 資料庫路徑。預設為 "stock_db.db"。

    """
    stock_list = [s.strip() for s in stocks.split(',')]
    print(f"[CLI] 開始更新股票資料: {stock_list}")
    with DBManager(db_path) as db_manager:
        for stock in stock_list:
            print(f"[CLI] 正在更新 {stock} 的資料...")
            db_manager.update_stock_in_db(stock)  # 更新資料庫中的股票資料
    print("更新完成。")

@app.command()
def index(stock: str, period: str = "1y", year: str = "2024", db_path: str = DB_PATH):
    """
    查詢股票技術指標資料(MACD)預設抓取 20, 50, 200，若資料庫沒有則從網路下載。
    """
    stock = stock.upper()
    db = DBManager(db_path)

    # 從資料庫抓資料（若無則自動 fallback 到 yfinance 並寫入 DB）
    df = db.get_stock(stock, period)

    if df.empty:
        typer.echo("⚠️ 沒有成功取得資料，請檢查股票代碼或網路連線。")
        return

    # 計算移動平均線
    newest = df["Close"].iloc[-1]
    ma20 = df["Close"].rolling(window=20).mean().iloc[-1]
    ma50 = df["Close"].rolling(window=50).mean().iloc[-1]
    ma200 = df["Close"].rolling(window=200).mean().iloc[-1]

    # 顯示
    typer.echo(f"\n📊 股票代碼：{stock}")
    typer.echo(f"最新收盤價: {newest:.2f}")
    typer.echo(f"目前 MA20: {ma20:.2f}")
    typer.echo(f"目前 MA50: {ma50:.2f}")
    typer.echo(f"目前 MA200: {ma200:.2f}")

    typer.echo(f"#########################################################")
    typer.echo("📈 Avg Pricing：")
    result = get_quarter_medians_and_year_avg(db_path, stock, year)

    typer.echo(f" {year} Q1 avg pricing: {result['Q1']}")
    typer.echo(f" {year} Q2 avg pricing: {result['Q2']}")
    typer.echo(f" {year} Q3 avg pricing: {result['Q3']}")
    typer.echo(f" {year} Q4 avg pricing: {result['Q4']}")
    typer.echo(f" {year} avg pricing: {result['YearAvg']}")

    db.close()


@app.command()
def list(db_path: str = DB_PATH):
    """
    列出資料庫中目前已儲存的所有股票代碼（根據資料表名開頭 t_ 推斷）。
    """
    with DBManager(db_path) as db_manager:
        stocks = db_manager.list_stocks()
        if stocks:
            print(f"[CLI] 資料庫目前有以下股票：{stocks}")
        else:
            print("[CLI] 資料庫中沒有任何股票。")


@app.command()
def autofill_excel(
    file: str,
    db_path: str = DB_PATH,
    target_stock: str = typer.Option(None, help="指定要處理的股票代碼"),
    target_year: int = typer.Option(None, help="指定要處理的年份")
):
    """
    自動填入 Excel 表格：從資料庫讀取每張工作表代表的股票資訊（最新價格、MA20/50/200、季中位數、年均價），
    並寫入對應 Excel 工作表內。
    """
    typer.echo(f"🚀 開始處理 Excel：{file}")
    wb = load_workbook(file)

    with DBManager(db_path) as db:
        for sheetname in wb.sheetnames:
            stock = sheetname.upper()
            if target_stock and stock != target_stock.upper():
                continue

            typer.echo(f"\n📄 處理工作表：{stock}")
            ws = wb[sheetname]

            # 抓價格與 MA
            df = db.get_stock(stock, period="1y")
            if df.empty:
                typer.echo(f"⚠️ 無法取得 {stock} 資料，略過。")
                continue

            close = df["Close"]
            newest = close.iloc[-1]
            ma20 = close.rolling(window=20).mean().iloc[-1]
            ma50 = close.rolling(window=50).mean().iloc[-1]
            ma200 = close.rolling(window=200).mean().iloc[-1]

            # 寫入欄位名稱與值
            ws["B1"] = "最新股價"
            ws["C1"] = "MA20"
            ws["D1"] = "MA50"
            ws["E1"] = "MA200"
            ws["B2"] = round(newest, 2)
            ws["C2"] = round(ma20, 2)
            ws["D2"] = round(ma50, 2)
            ws["E2"] = round(ma200, 2)

            # 找指定年份的每一季度列位置，從該行 F 欄開始填入
            for i in range(3, ws.max_row + 1):
                year_val = ws[f"A{i}"].value
                type_val = ws[f"B{i}"].value

                if type_val == "Y" and isinstance(year_val, int):
                    year = year_val
                    if target_year and year != target_year:
                        continue

                    typer.echo(f"  ↳ 處理年度：{year}")
                    result = get_quarter_medians_and_year_avg(db_path, stock, year)

                    # 從 Y 那一列，找接下來四列 (Q4, Q3, Q2, Q1)
                    quarters = ["Q4", "Q3", "Q2", "Q1"]
                    for offset, q in enumerate(quarters):
                        row_idx = i + offset + 1
                        median_price = result.get(q)
                        if median_price:
                            ws[f"F{row_idx}"] = median_price

                    # Y 年那一列放入年均價
                    ws[f"F{i}"] = result.get("YearAvg")

    # 儲存於原始檔案
    wb.save(file)
    typer.echo(f"\n✅ 已更新原始檔案：{file}")

@app.command()
def market(stock: str):
    """
    爬取 finviz 資料，蒐集如當前 EPS, ROE 等資訊。
    """
    data = get_finviz_data(stock)

    print(f"📊 {stock} 財報摘要：")
    print(f"EPS (ttm): {data.get('EPS (ttm)')}")
    print(f"ROE: {data.get('ROE')}")
    print(f"Gross Margin: {data.get('Gross Margin')}")
    print(f"Debt/Eq: {data.get('Debt/Eq')}")
    print(f"P/E: {data.get('P/E')}")
    print(f"Forward P/E: {data.get('Forward P/E')}")

if __name__ == "__main__":
    app()